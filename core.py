"""
Lógica pura do dashboard FIC FIDC: parsing + agregação de movimentações por
pessoa + geração do texto. Sem dependência de Streamlit.

Não há classificação de cotista/prestador. Toda contraparte que aparece em
movimentações PIX/TED é agregada e listada com entrada, saída e líquido.
O saldo (fundo de liquidez + conta) é apresentado à parte.
"""
from __future__ import annotations

import re
from io import BytesIO
from typing import IO

import pandas as pd
from openpyxl import load_workbook

from config import SUFIXOS_EMPRESA


# =============================================================================
# Formatação
# =============================================================================

def fmt_brl(v: float) -> str:
    """Formata número como 'R$ 1.234,56' (padrão brasileiro)."""
    if v == 0:  # normaliza -0.0
        v = 0.0
    s = f"{v:,.2f}"
    return "R$ " + s.replace(",", "X").replace(".", ",").replace("X", ".")


# =============================================================================
# Normalização de nomes de contraparte
# =============================================================================

def normaliza_parte(comp: str) -> str:
    """Remove prefixos 'Dest:'/'Rem:', sufixo '-ESTORNO', colapsa espaços, MAIÚSCULAS."""
    s = str(comp or "").strip()
    s = re.sub(r"^(Dest:|Rem:)\s*", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s*-\s*ESTORNO\s*$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s).strip()
    return s.upper()


def chave_parte(nome_norm: str) -> str:
    """
    Chave de agrupamento: remove sufixos societários (LTDA, S.A., ...) para juntar
    'ELANTRA PARTICIPACOES' e 'ELANTRA PARTICIPACOES LTDA' como a mesma pessoa.
    """
    s = nome_norm
    mudou = True
    while mudou:
        mudou = False
        for suf in SUFIXOS_EMPRESA:
            padrao = r"\s*\b" + re.escape(suf) + r"\b\.?\s*$"
            novo = re.sub(padrao, "", s, flags=re.IGNORECASE).strip()
            if novo != s:
                s = novo
                mudou = True
    return re.sub(r"[.\s]+$", "", s).strip()


def nome_exibicao(nome_norm: str) -> str:
    """'KOVR CAPITALIZACAO S A' → 'Kovr Capitalizacao' (limpa sufixo + title case)."""
    return chave_parte(nome_norm).title()


# =============================================================================
# Parsers
# =============================================================================

def _ler_csv_bytes(data: bytes) -> pd.DataFrame:
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            return pd.read_csv(BytesIO(data), sep=";", encoding=enc)
        except UnicodeDecodeError:
            continue
    raise ValueError("Não foi possível decodificar o CSV.")


def parse_extrato(file: IO[bytes] | str) -> pd.DataFrame:
    """Lê o extrato (CSV) e cria `valor_assinado` (+ crédito, − débito)."""
    if isinstance(file, str):
        with open(file, "rb") as f:
            data = f.read()
    else:
        data = file.read()
    df = _ler_csv_bytes(data)

    df["dt_mov"] = pd.to_datetime(df["dt_mov"], format="%d/%m/%Y")
    df["valor"] = pd.to_numeric(df["valor"])
    df["valor_assinado"] = df.apply(
        lambda r: r["valor"] if r["fl_debito_credito"] == "C" else -r["valor"],
        axis=1,
    )
    for col in ("ds_historico", "ds_complemento"):
        df[col] = df[col].fillna("").astype(str)
    return df.sort_values("dt_mov").reset_index(drop=True)


def parse_carteira(file: IO[bytes] | str) -> dict:
    """Lê um XLSX 'Saldo de Aplicações'. Retorna {saldo_liquido, fundo_liquidez, cotista_nome}."""
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    saldo_col = None
    fundo_liquidez = None
    cotista_nome = None
    saldo_liquido = None

    for row in rows:
        if not row:
            continue
        for j, cell in enumerate(row):
            if cell is None:
                continue
            s = str(cell).strip()

            if saldo_col is None and ("Saldo Líquido" in s or "Saldo Liquido" in s):
                saldo_col = j

            if fundo_liquidez is None and "Carteira:" in s:
                for k in range(j + 1, len(row)):
                    nxt = row[k]
                    if nxt is None:
                        continue
                    nxt_s = str(nxt).strip()
                    if nxt_s and "CNPJ" not in nxt_s:
                        fundo_liquidez = nxt_s
                        break

            if cotista_nome is None:
                m = re.match(r"^\s*\d+\s*-\s*(.+?)\s*-\s*(Aberto|Fechado)", s)
                if m:
                    cotista_nome = m.group(1).strip()

            if saldo_liquido is None and s.startswith("Total:") and saldo_col is not None:
                if len(row) > saldo_col and row[saldo_col] is not None:
                    try:
                        saldo_liquido = float(row[saldo_col])
                    except (ValueError, TypeError):
                        pass

    return {
        "saldo_liquido": saldo_liquido,
        "fundo_liquidez": fundo_liquidez,
        "cotista_nome": cotista_nome,
    }


# =============================================================================
# Agregação e saldo
# =============================================================================

def computar_saldo_conta(df: pd.DataFrame) -> float:
    """Saldo final em conta-corrente = soma dos valores assinados."""
    return float(df["valor_assinado"].sum())


def somar_saldo_liquidez(carteiras: list[dict]) -> float:
    """Soma o saldo_liquido de várias carteiras (ignora None)."""
    return float(sum(c.get("saldo_liquido") or 0.0 for c in carteiras))


def agregar_pessoas(df: pd.DataFrame) -> pd.DataFrame:
    """
    Agrega cada contraparte das movimentações PIX/TED por nome (variações do
    mesmo nome são unidas).

    Entrada = créditos (dinheiro que entrou na conta do fundo);
    Saída   = débitos (dinheiro que saiu, inclui devoluções/estornos).
    A classificação usa o flag débito/crédito do extrato — assim uma linha
    'PIX - RECEBIDO DEVOLVIDO' (que é um débito) entra corretamente na saída.

    Retorna DataFrame [nome, entrada, saida, liquido] ordenado por líquido desc.
    Linhas sem entrada e sem saída (ex.: tarifas) são descartadas.
    """
    movs = df[
        df["ds_historico"].str.contains(r"PIX|TED|TEC", case=False, regex=True, na=False)
        & ~df["ds_historico"].str.contains(r"TARIFA", case=False, regex=True, na=False)
    ].copy()

    if movs.empty:
        return pd.DataFrame(columns=["nome", "entrada", "saida", "liquido"])

    movs["parte_norm"] = movs["ds_complemento"].apply(normaliza_parte)
    movs["chave"] = movs["parte_norm"].apply(chave_parte)
    movs["credito"] = movs["fl_debito_credito"] == "C"

    registros = []
    for chave, grupo in movs.groupby("chave"):
        if not chave:
            continue
        entrada = float(grupo.loc[grupo["credito"], "valor"].sum())
        saida = float(grupo.loc[~grupo["credito"], "valor"].sum())
        if entrada == 0 and saida == 0:
            continue
        nome_norm = max(grupo["parte_norm"], key=len)
        registros.append({
            "nome": nome_exibicao(nome_norm),
            "entrada": entrada,
            "saida": saida,
            "liquido": entrada - saida,
        })

    out = pd.DataFrame(registros)
    if out.empty:
        return out
    return out.sort_values("liquido", ascending=False).reset_index(drop=True)


# =============================================================================
# Geração do texto para WhatsApp
# =============================================================================

def gerar_resumo(nome_fundo: str, saldo_total: float, pessoas: pd.DataFrame) -> str:
    """Monta o texto: cabeçalho, lista por pessoa e, à parte, o saldo."""
    out = [nome_fundo, ""]

    for _, r in pessoas.sort_values("liquido", ascending=False).iterrows():
        out.append(f"{r['nome']}:")
        out.append(f"Entrada: {fmt_brl(r['entrada'])}")
        out.append(f"Saída: {fmt_brl(r['saida'])}")
        out.append(f"Líquido: {fmt_brl(r['liquido'])}")

    out.append("")
    out.append(f"Saldo (fundo de liquidez + conta): {fmt_brl(saldo_total)}")
    return "\n".join(out)


def limpar_nome_fundo(nome_bruto: str | None) -> str:
    """'FIC FIDC ELANTRA RL' → 'FIC FIDC Elantra'."""
    if not nome_bruto:
        return "FIC FIDC"
    s = nome_bruto.title()
    s = re.sub(r"\bFic Fidc\b", "FIC FIDC", s, flags=re.IGNORECASE)
    s = re.sub(r"\bFidc\b", "FIDC", s, flags=re.IGNORECASE)
    s = re.sub(r"\bRl\b\s*$", "", s).strip()
    return s
